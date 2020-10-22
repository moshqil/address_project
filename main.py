import pandas as pd
import googlemaps
import datetime
import time
import pprint
import pickle
import string
API_KEY = ...
gmaps = googlemaps.Client(key=API_KEY)
data = pd.read_csv('address.csv', sep=';')

student_names = []
address = []
for i, info in data.iterrows():
    address.append(info['address'])
    student_names.append(info['name'])

distance_matrix = [[-1] * 26 for _ in range(26)]
departure_time = datetime.datetime(2018, 11, 29, 12)
for i in range(1, 26):
    for j in range(0, 26):
        if i == j:
            continue
        distance_matrix[i][j] = gmaps.directions(address[i], address[j], mode="transit", departure_time=departure_time)
        print(distance_matrix[i][j][0]['legs'][0]['duration'])
        time.sleep(1)
i = 23
for j in range(0, 26):
    if i == j:
        continue
    distance_matrix[i][j] = gmaps.directions(address[i], address[j], mode="transit", departure_time=departure_time)
    distance_matrix[j][i] = gmaps.directions(address[j], address[i], mode="transit", departure_time=departure_time)
    print(distance_matrix[i][j][0]['legs'][0]['duration'])
    time.sleep(1)
distance_matrix[23][25] = gmaps.directions(address[23], address[25], mode="transit", arrival_time=datetime.datetime(2018, 12, 3, 9))
dist_time_m = [[distance_matrix[i][j][0]['legs'][0]['duration']['value'] if i != j else 0 for j in range(26)] for i in range(26)]

# Среднее время от всех до всех

all_time = sum(sum(row) for row in dist_time_m)
all_time / (26 * 25 / 2) / 60

# Среднее время на путь в школу

print(sum(dist_time_m[i][25] + dist_time_m[25][i] for i in range(25)) / 50 / 60)

# Максимальное и минимальное время в школу:

print(max((dist_time_m[i][25] / 60, i) for i in range(24)))

f = open('dist.pickle', 'rb')
dm = distance_matrix = pickle.load(f)
f = open('dist.pickle', 'wb')
pickle.dump(distance_matrix, f)
distance_matrix[0][6] = gmaps.directions(address[0], address[6], mode="transit", departure_time=departure_time)

from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws.title = 'Данные'

names = list(string.ascii_uppercase) + ['AA', 'AB']
for i in range(26):
    for j in range(26):
        if i == j:
            continue
        #print(i,j)
        ij = distance_matrix[i][j][0]['legs'][0]['duration']['value']
        ws[names[i + 2] + str(j + 3)] = f'{ij // 3600}ч, {ij % 3600 // 60}м'
        ji = distance_matrix[j][i][0]['legs'][0]['duration']['value']
        #print(ji)
        ws[names[j + 2] + str(i + 3)] = f'{ji // 3600}ч, {ji % 3600 // 60}м'

for i in range(26):
    ws['B' + str(i + 3)] = student_names[i]
for i in range(26):
    ws[names[i + 2] + '2'] = student_names[i]

for i in range(28):
    ws[names[i] + str(i + 1)] = ''
    ws[names[i] + '1'] = ''
    ws['A' + str(i + 1)] = ''

ws['C1'] = 'Откуда'
ws['A3'] = 'Куда'

wb.save('time.xlsx')

sch57 = 'Малый Знаменский переулок, 7/10 стр. 5'
to_sch57 = [0] * 25
for i in range(25):
    to_sch57[i] = gmaps.directions(address[i], sch57, mode="transit", arrival_time=datetime.datetime(2018, 12, 3, 9))
    time.sleep(1)
 print((sum(to_sch57[i][0]['legs'][0]['duration']['value'] for i in range(25) if to_sch57[i] != []) / 22 / 60))
 print(([distance_matrix[i][25][0]['legs'][0]['duration']['value'] / 60 for i in range(25) if to_sch57[i] != []]))
